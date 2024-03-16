from selenium.webdriver.chrome.service import Service
from selenium import webdriver
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

# path = "E:/dev/python/RPA-Challenge/xlsx/challenge.xlsx"
path = "./xlsx/challenge.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

row_count = 0
for row in sheet_obj:
    if not all([cell.value == None for cell in row]):
        row_count += 1

max_col = sheet_obj.max_column

xlsx_data_list = []
for i in range(2, row_count + 1):
    for j in range(1, max_col):
        cell_obj = sheet_obj.cell(row = i, column= j)
        xlsx_data_list.append(cell_obj.value)


def list_creation(xlsx_data_list):
    input_list = [] 
    idx = 0
    for i in range(row_count -1): # 10
        for j in range(max_col- 1): # 7
            if idx == len(xlsx_data_list):
                break
            else:
                input_list.append(xlsx_data_list[idx])
                idx += 1
        data_input(input_list)


chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

# service = Service(executable_path='E:/dev/python/RPA-Challenge/chromedriver-win64/chromedriver.exe')
service = Service(executable_path='./chromedriver-win64/chromedriver.exe')
driver = webdriver.Chrome(service=service, options=chrome_options)
driver.maximize_window()

rpa_website = driver.get("http://www.rpachallenge.com/")

def data_input(input_list):
    start_button = driver.find_element(By.CLASS_NAME, "btn-large.uiColorButton")
    count = 0
    while count < 1:
        start_button.click()
        count += 1
    first_name_form = driver.find_element(By.XPATH,'//*[@ng-reflect-name="labelFirstName"]')
    last_name_form = driver.find_element(By.XPATH,'//*[@ng-reflect-name="labelLastName"]')
    cmpny_name_form = driver.find_element(By.XPATH,'//*[@ng-reflect-name="labelCompanyName"]')
    role_in_cmpny_form = driver.find_element(By.XPATH,'//*[@ng-reflect-name="labelRole"]')
    address_form = driver.find_element(By.XPATH,'//*[@ng-reflect-name="labelAddress"]')
    email_form = driver.find_element(By.XPATH,'//*[@ng-reflect-name="labelEmail"]')
    phone_num_form = driver.find_element(By.XPATH,'//*[@ng-reflect-name="labelPhone"]')
    submit_button = driver.find_element(By.CLASS_NAME,'btn.uiColorButton')
    
    first_name_form.click()
    first_name_form.send_keys(input_list[0]) # 0 = First Name 
    last_name_form.click()
    last_name_form.send_keys(input_list[1]) # 1 = Last Name 
    cmpny_name_form.click()
    cmpny_name_form.send_keys(input_list[2]) # 2 = Company Name 
    role_in_cmpny_form.click()
    role_in_cmpny_form.send_keys(input_list[3]) # 3 = Role in Company 
    address_form.click()
    address_form.send_keys(input_list[4]) # 4 = Address
    email_form.click()
    email_form.send_keys(input_list[5]) # 5 = Email
    phone_num_form.click()
    phone_num_form.send_keys(input_list[6]) # 6 = Phone Number 
    submit_button.click()
    input_list.clear()
    return

list_creation(xlsx_data_list)